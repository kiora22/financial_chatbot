"""
Microsoft Excel spreadsheet processor.
Handles .xlsx files using pandas and openpyxl.
"""

import os
from typing import Dict, Any, Optional
import pandas as pd
import openpyxl

from document_processor.processors.base_processor import BaseDocumentProcessor
from config.logging_config import setup_logging

# Set up logging
logger = setup_logging("xlsx_processor")


class ExcelProcessor(BaseDocumentProcessor):
    """Processor for Microsoft Excel spreadsheets (.xlsx)."""
    
    async def extract_text(self, file_path: str) -> str:
        """
        Extract text from an Excel spreadsheet.
        
        Args:
            file_path: Path to the Excel file
            
        Returns:
            Extracted text
        """
        try:
            logger.info(f"Extracting text from Excel spreadsheet: {file_path}")
            
            # Read all sheets into a dictionary of dataframes
            xl = pd.ExcelFile(file_path)
            sheet_names = xl.sheet_names
            
            # Extract text from each sheet
            sheets_text = []
            
            for sheet_name in sheet_names:
                logger.debug(f"Processing sheet: {sheet_name}")
                
                # Read the sheet into a dataframe
                df = pd.read_excel(file_path, sheet_name=sheet_name)
                
                # Skip empty dataframes
                if df.empty:
                    continue
                
                # Convert to string representation
                sheet_text = f"--- Sheet: {sheet_name} ---\n\n"
                
                # Get column names
                columns_text = "Columns: " + ", ".join(str(col) for col in df.columns) + "\n\n"
                sheet_text += columns_text
                
                # Convert dataframe to string representation
                data_text = df.to_string(index=False, na_rep="")
                sheet_text += data_text + "\n\n"
                
                sheets_text.append(sheet_text)
            
            # Join all sheets
            text = "\n\n".join(sheets_text)
            
            logger.info(f"Successfully extracted {len(text)} characters from {file_path}")
            return text
            
        except Exception as e:
            logger.error(f"Error extracting text from Excel spreadsheet {file_path}: {str(e)}")
            return ""
    
    async def extract_metadata(self, file_path: str) -> Optional[Dict[str, Any]]:
        """
        Extract metadata from an Excel spreadsheet.
        
        Args:
            file_path: Path to the Excel file
            
        Returns:
            Extracted metadata
        """
        try:
            # Get file stats
            stat_info = os.stat(file_path)
            
            # Initialize metadata with file stats
            metadata = {
                "file_size_bytes": stat_info.st_size,
                "created_at": stat_info.st_ctime,
                "modified_at": stat_info.st_mtime,
                "file_extension": "xlsx"
            }
            
            # Load workbook to extract Excel-specific metadata
            xl = pd.ExcelFile(file_path)
            metadata["sheet_names"] = xl.sheet_names
            metadata["sheet_count"] = len(xl.sheet_names)
            
            # Get more detailed properties using openpyxl
            try:
                wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
                
                # Get properties if available
                if wb.properties:
                    props = wb.properties
                    if props.creator:
                        metadata["creator"] = props.creator
                    if props.title:
                        metadata["title"] = props.title
                    if props.subject:
                        metadata["subject"] = props.subject
                    if props.created:
                        metadata["doc_created_at"] = props.created.isoformat()
                    if props.modified:
                        metadata["doc_modified_at"] = props.modified.isoformat()
                
                # Get sheet info
                sheet_info = []
                for sheet_name in wb.sheetnames:
                    sheet = wb[sheet_name]
                    sheet_info.append({
                        "name": sheet_name,
                        "max_row": sheet.max_row,
                        "max_column": sheet.max_column
                    })
                
                metadata["detailed_sheet_info"] = sheet_info
            except Exception as e:
                logger.warning(f"Error getting detailed Excel metadata: {str(e)}")
            
            return metadata
            
        except Exception as e:
            logger.error(f"Error extracting metadata from Excel spreadsheet {file_path}: {str(e)}")
            return None